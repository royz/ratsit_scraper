import re
from pprint import pprint
import json
import requests
from bs4 import BeautifulSoup
import os
import glob
import openpyxl
import time
import colorama
from termcolor import colored
from requests.exceptions import ReadTimeout

BASE_DIR = os.path.dirname(__file__)
CACHE_PATH = os.path.join(BASE_DIR, 'cache.json')


class Ratsit:
    def __init__(self):
        self.session = None
        self.cache = None
        self.cache_written_at = time.time()
        self.read_cache()
        self.init_session()

    def init_session(self):
        self.session = requests.session()
        self.session.headers = {
            'authority': 'www.ratsit.se',
            'accept': '*/*',
            'x-requested-with': 'XMLHttpRequest',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/88.0.4324.182 Safari/537.36',
            'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'origin': 'https://www.ratsit.se',
        }

    def search(self, first_name, last_name, person_number):
        if not self.cache:
            self.read_cache()

        person_hash = self.get_hash(first_name, last_name, person_number)

        if person_hash in self.cache:
            return self.cache.get(person_hash)

        params = {
            'fnamn': first_name,
            'enamn': last_name,
            'gata': '',
            'postnr': '',
            'ort': '',
            'kn': '',
            'pnr': person_number,
            'tfn': '',
            'm': '1',
            'k': '1',
            'r': '1',
            'er': '1',
            'b': '1',
            'eb': '1',
            'amin': '16',
            'amax': '120',
            'fon': '1',
            'typ': '2',
            'page': '1',
        }
        try:
            response = self.session.get('https://www.ratsit.se/api/search/person', params=params, timeout=10)
        except ReadTimeout:
            print('request timed out (2). retrying...')
            return self.search(first_name, last_name, person_number)

        try:
            url = 'https://www.ratsit.se' + response.json()['person']['list'][0]['personrapportUrl']
            return self.get_details(url, person_hash)
        except:
            return None

    def get_details(self, url, person_hash):
        try:
            response = requests.get(url, timeout=10)
        except ReadTimeout:
            print('request timed out (1). retrying...')
            return self.get_details(url, person_hash)

        # with open('details.html', 'w', encoding='utf-8') as f:
        #     f.write(response.text)

        soup = BeautifulSoup(response.text, 'html.parser')

        # get the address and phone number
        data = self.find_address(soup)
        data.update({'url': url})

        # get all company names
        try:
            company_spans = soup.find_all('span', {'class': 'engagement-company'})
            companies = [company_span.text.strip() for company_span in company_spans]
            companies = list(set(companies))
        except:
            companies = []
        data.update({'companies': companies})

        # get businesses
        try:
            business_table = soup.find('div', {'id': 'foretagPaAdressenLista'})
            table_rows = business_table.find_all('tr')[1:]
            businesses = [business.find('a').text.strip() for business in table_rows]
        except Exception as e:
            businesses = []

        data.update({'businesses': businesses})

        # get a list of persons they are living with
        persons_living_with = []
        try:
            report_table = soup.find('table', {'class': 'rapport-table rapport-table--limit-large-screens'})
            for tr in report_table.find_all('tr'):
                persons_living_with.append(tr.text.strip().replace('\n', ''))
        except:
            pass
        data.update({'living_with': persons_living_with})

        # add the data into cache
        self.cache[person_hash] = data
        # save the cache into a file
        self.write_cache()

        return data

    @staticmethod
    def print_details(details):
        address = details.get('address')
        phone = details.get('phone')
        companies = details.get('companies')
        businesses = details.get('businesses')
        living_with = details.get('living_with')

        if address:
            print(colored(f"address: {address['street']}, {address['locality']}, {address['zip']}", 'green'))
        else:
            print(colored('address: not found', 'yellow'))

        if len(living_with) > 0:
            print(colored(f'living with: {join_list(living_with)}', 'green'))
        else:
            print(colored('living with: not found', 'yellow'))

        if phone:
            print(colored(f"phone: {phone}", 'green'))
        else:
            print(colored('phone: not found', 'yellow'))

        if len(companies) > 0:
            print(colored(f"companies: {join_list(companies)}", 'green'))
        else:
            print(colored('company: not found', 'yellow'))

        if len(businesses) > 0:
            print(colored(f"businesses: {join_list(businesses)}", 'green'))
        else:
            print(colored('business: not found', 'yellow'))

    @staticmethod
    def find_address(soup):
        new_data = {'address': None, 'phone': None}
        try:
            data = soup.find('script', {'type': 'application/ld+json'})
            json_data = json.loads(next(data.children).strip())
            for res in json_data:
                # if address already not found and address is present in current data
                if not new_data['address'] and 'address' in res:
                    # get the address
                    try:
                        new_data['address'] = {
                            'country': res['address']['addressCountry'],
                            'locality': res['address']['addressLocality'],
                            'zip': res['address']['postalCode'],
                            'street': res['address']['streetAddress']
                        }
                    except:
                        pass

                    # get the phone number
                    try:
                        new_data['phone'] = res['telephone']
                    except:
                        pass
        except:
            pass

        return new_data

    @staticmethod
    def get_hash(first_name, last_name, person_number):
        """
        create an unique hashable string with the person's info
        :param first_name: str
        :param last_name: str
        :param person_number: str
        :return: str
        """
        return f'{first_name}-{last_name}-{person_number}'

    def read_cache(self):

        if os.path.exists(CACHE_PATH):
            try:
                with open(CACHE_PATH, encoding='utf-8') as f:
                    self.cache = json.load(f)
            except:
                self.cache = {}
        else:
            self.cache = {}

    def write_cache(self):
        # write the cache every 15 seconds
        if time.time() < self.cache_written_at + 15:
            return

        with open(CACHE_PATH, 'w', encoding='utf-8') as f:
            json.dump(self.cache, f, indent=2)


class Cell:
    address = 4
    zip = 5
    city = 6
    phone = 7
    info = 10


class Excel:
    def __init__(self):
        self.file_name = None
        self.file_path = None
        self.input = []
        self.output = []

    def read_input(self):
        files = glob.glob(os.path.join(BASE_DIR, '*.xlsx'))
        file_names = [file.split(os.path.sep)[-1] for file in files]
        print('choose a file:')
        for i, file in enumerate(file_names):
            print(f'{i + 1}. {file}')
        try:
            idx = int(input('file number: ')) - 1
            self.file_name = file_names[idx]
            self.file_path = files[idx]

            # read the data
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            for i, row in enumerate(sheet.iter_rows()):
                if i == 0:
                    continue

                if not row[1].value and not row[2].value:
                    break

                self.input.append({
                    'row': i,
                    'first_name': row[1].value,
                    'last_name': row[2].value,
                    'person_number': row[0].value,
                })
        except StopIteration:
            print('the file does not exist')
            quit()

    def write_data(self):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active

        for row in self.output:
            info = ''
            if len(row['living_with']) > 0:
                info += f'living with: {join_list(row["living_with"])}\n'
            if len(row['companies']) > 0:
                info += f'companies: {join_list(row["companies"])}\n'
            if len(row['businesses']) > 0:
                info += f'businesses: {join_list(row["businesses"])}'

            sheet.cell(row=row['row'], column=Cell.info).value = info.strip()

            if row.get('address'):
                sheet.cell(row=row['row'], column=Cell.address).value = row['address']['street']
                sheet.cell(row=row['row'], column=Cell.zip).value = row['address']['zip']
                sheet.cell(row=row['row'], column=Cell.city).value = row['address']['locality']
            if row.get('phone'):
                sheet.cell(row=row['row'], column=Cell.phone).value = row['phone']

        workbook.save(self.file_path)
        print(colored(f'data saved in {self.file_name}', 'green'))


def join_list(_list: list) -> str:
    _list = [item if ',' not in item else f'"{item}"' for item in _list]
    _list = list(map(lambda x: x.replace('\n', '').replace('  ', ' '), _list))
    return ', '.join(_list)


if __name__ == '__main__':
    colorama.init()

    ratsit = Ratsit()
    excel = Excel()
    excel.read_input()

    for i, person in enumerate(excel.input):
        if not person["first_name"] and not person["last_name"] and not person["person_number"]:
            break

        print(colored(f'[{i + 1}/{len(excel.input)}] {person["first_name"]} '
                      f'{person["last_name"]} ({person["person_number"]}) '
                      f''.ljust(60, '-'), 'blue'))
        result = ratsit.search(
            person['first_name'],
            person['last_name'],
            person['person_number'],
        )
        if not result:
            print(colored('no results found', 'red'))
            continue

        output = result
        output.update({'row': person['row'] + 1})

        ratsit.print_details(result)

        # add the data to be written to excel
        excel.output.append(output)

    # write the data into cache
    ratsit.write_cache()

    # save the data to the excel file
    excel.write_data()
